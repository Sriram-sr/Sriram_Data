import openpyxl

excel_file_path  = '/home/sriram/Selenium/Fexc.xlsx'

workbook = openpyxl.load_workbook(excel_file_path)

sheets = workbook.sheetnames  #This is a list

current_sheet = workbook[sheets[1]]  #Handling different sheets

no_of_rows = current_sheet.max_row
no_of_cols = current_sheet.max_column

for row in range(1,no_of_rows+1):
    for col in range(1,no_of_cols+1):
        print(current_sheet.cell(row=row,column=col).value,end = '     ')
    print()  